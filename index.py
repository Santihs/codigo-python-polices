from tkinter import *
from tkinter import ttk
from tkinter.font import Font
from tkinter.ttk import Combobox
from xml.etree.ElementTree import Element
from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import messagebox

class Interfaz:

    def __init__(self, ventana):
        self.ventana = ventana
        self.ventana.title('Buscador en excel')
        # self.ventana.iconbitmap('./img/police2.ico')
        self.ventana.geometry("1040x550+100+50")

        # creamos un frame
        frame = LabelFrame(self.ventana,text="Buscador")
        frame.grid(row=0,column=0,pady=2)

        # creamos el label y el buscador
        self.texto = Entry(frame, width=46) 
        self.texto.focus()
        self.texto.grid(row=0,column=0,padx=3,pady=3)

        # creamos el boton
        ttk.Button(frame,text="Buscar", command=self.buscar).grid(row=0,column=1,padx=0,pady=3)
        ttk.Button(frame,text="Refrescar", command=self.abrir_Excel).grid(row=0,column=2,padx=3,pady=3)
        # ttk.Button(frame,text="Abrir...", command=self.abreArchivo).grid(row=3,column=1,padx=3,pady=3)
        ttk.Button(frame,text="Ver", command=self.ver).grid(row=3,column=1,padx=3,pady=3)
        ttk.Button(frame,text="Editar", command=self.editar).grid(row=3,column=2,padx=3,pady=3)
        ttk.Button(frame,text="Añadir", command=self.aniadir).grid(row=3,column=3,padx=3,pady=3)
        ttk.Button(frame,text="Eliminar", command=self.eliminar).grid(row=3,column=4,padx=3,pady=3)

        # creamos una tabla
        self.tree = ttk.Treeview(frame,height=22,columns=7)
        self.tree.grid(row=1,column=0,columnspan=5)

        self.tree["columns"] = ("1","2","3","4","5","6","7")

        self.tree.column("#0", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("1", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("2", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("3", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("4", minwidth=160, width=150,anchor=CENTER)
        self.tree.column("5", minwidth=160, width=160,anchor=CENTER)
        self.tree.column("6", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("7", minwidth=120, width=120,anchor=CENTER)

        self.tree.heading("#0",text="C.I.")
        self.tree.heading("1",text="Apellido Paterno")
        self.tree.heading("2",text="Apellido Materno")
        self.tree.heading("3",text="Nombres")
        self.tree.heading("4",text="Grado")
        self.tree.heading("5",text="Lugar de Nacimiento")
        self.tree.heading("6",text="Fecha de Nacimiento")
        self.tree.heading("7",text="Escalafón")

        self.abrir_Excel()
        # self.tree.insert("","end",text="Huanacu",values=("Sanchez","santiago  jose","coronel"))

    def limpiar(self):
        records = self.tree.get_children()
        for element in records:
            self.tree.delete(element)
        return

    def abrir_Excel(self):
        self.limpiar()

        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        sheet = wb[nombresHojas[0]]

        self.listaDeListas =[]

        for fila in range(2,sheet.max_row+1):
            lista = []
            for columna in range(1,sheet.max_column+1):
                lista.append(str(sheet.cell(row=fila,column=columna).value))
            self.listaDeListas.append(lista)

        for fila in range(0,len(self.listaDeListas)):
            self.tree.insert("","end",text=self.listaDeListas[fila][0],
                            values=(self.listaDeListas[fila][1],self.listaDeListas[fila][2],
                                    self.listaDeListas[fila][3],self.listaDeListas[fila][4],
                                    self.listaDeListas[fila][5],self.listaDeListas[fila][6],
                                    self.listaDeListas[fila][7]))
        return

    def ver(self):
        try:
            self.tree.item(self.tree.selection())['text'][0]
        except IndexError as e:
            messagebox.showwarning("Advertencia", "Seleccione un dato")
            return
        ci = str(self.tree.item(self.tree.selection())['text'])
        pos = self.buscar_pos(ci)
        listaMostrar = self.listaDeListas[pos]

        self.ventanaVer = Ver(self.ventana,listaMostrar,pos)
    
    def eliminar(self):
        valor = messagebox.askokcancel("Eliminar", "¿Esta seguro que desea eliminarlo?")
        if valor == True:
            self.eliminarSeleccion()
        
    def eliminarSeleccion(self):
        try:
            self.tree.item(self.tree.selection())['text'][0]
        except IndexError as e:
            messagebox.showwarning("Advertencia", "Seleccione un dato")
            return
        ci = str(self.tree.item(self.tree.selection())['text'])
        pos = self.buscar_pos(ci)
        listaMostrar = self.listaDeListas[pos]

        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        for nroHoja in range(3):
            sheet = wb[nombresHojas[nroHoja]]
            self.eliminarFila(sheet,listaMostrar[0],wb)
        self.abrir_Excel()

    def eliminarFila(self,sheet,ci,wb):
        res = []
        for fila in range(2,sheet.max_row+1):
            for columna in range(1,sheet.max_column+1):
                if str(ci) in (str(sheet.cell(row=fila,column=columna).value)):
                    res.append(fila)
                    break
        for i in res:        
            sheet.delete_rows(i)
        wb.save(self.filesheet)
    
    def editar(self):
        try:
            self.tree.item(self.tree.selection())['text'][0]
        except IndexError as e:
            messagebox.showwarning("Advertencia", "Seleccione un dato")
            return
        ci = str(self.tree.item(self.tree.selection())['text'])
        pos = self.buscar_pos(ci)
        listaMostrar = self.listaDeListas[pos]

        self.ventanaEditar = Editar(self.ventana,listaMostrar,pos)
        self.abrir_Excel()

    def aniadir(self):
        self.ventanaAniadir = Aniadir(self.ventana)

    def buscar_pos(self, ci):
        pos = -1
        listapos = []
        for lista in self.listaDeListas:
            pos+=1
            for dato in lista:
                if ci.lower() in dato:
                    listapos.append(pos)
                    break
        res = listapos[0]
        return res

    def validation(self):
        return len(self.texto.get())>0

    def buscar(self):
        pos = -1
        listapos = []
        buscar = self.texto.get()
        # lista_res = self.convertir_Str()
        if self.validation():
            for lista in self.listaDeListas:
                pos+=1
                for dato in lista:
                    if buscar in dato.lower():
                        listapos.append(pos)
                        break
        elif self.validation()==None or self.validation()==0:
            self.abrir_Excel()
        self.texto.delete(0,END)
        self.mostrar_buscados(listapos)

    def mostrar_buscados(self, listapos):
        self.limpiar()
        for fila in listapos:
            self.tree.insert("","end",text=self.listaDeListas[fila][0],
                            values=(self.listaDeListas[fila][1],self.listaDeListas[fila][2],
                                    self.listaDeListas[fila][3],self.listaDeListas[fila][4],
                                    self.listaDeListas[fila][5],self.listaDeListas[fila][6],
                                    self.listaDeListas[fila][7]))
        return

class Ver:
    def __init__(self, ventana,listaMostrar, pos):
        self.newWindow = Toplevel(ventana)
        self.newWindow.title('Datos Personal')
        # self.newWindow.iconbitmap('./img/police2.ico')
        self.newWindow.geometry("770x570+150+100")

        self.miCanvas = Canvas(self.newWindow,background="white", width=770,height=570)
        framePrincipal = Frame(self.miCanvas)
        miScroll = Scrollbar(self.newWindow, orient="vertical")        

        miScroll.pack(side="right",fill="y")
        miScroll.config(command=self.miCanvas.yview)
        self.miCanvas.configure(yscrollcommand=miScroll.set)
        self.miCanvas.pack(side="left",expand=True,fill="both")
        self.miCanvas.create_window((0,0),window=framePrincipal,anchor="nw")

        self.cargar_ventanas(framePrincipal,listaMostrar, pos)
        
        self.newWindow.update()
        self.miCanvas.config(scrollregion=self.miCanvas.bbox("all"))

    def cargar_ventanas(self, framePrincipal, listaMostrar,pos):
        self.lista_mostrar = listaMostrar

        frame = LabelFrame(framePrincipal,text="Datos Personales del Funcionario", font="Arial 10 bold")
        frame.grid(row=0,column=0,pady=2)
        frame.configure(bg="#fff")

        Label(frame, text=listaMostrar[1], bg="white").grid(row=0,column=0,padx=5,pady=3)
        Label(frame, text=listaMostrar[2], bg="white").grid(row=0,column=1,padx=5,pady=3)
        Label(frame, text=listaMostrar[3], bg="white").grid(row=0,column=2,padx=5,pady=3)
        Label(frame, text=listaMostrar[4], bg="white").grid(row=0,column=3,padx=5,pady=3)
        Label(frame, text="Apellido Paterno", font="Arial 9 bold").grid(row=1,column=0,padx=20)
        Label(frame, text="Apellido Materno", font="Arial 9 bold").grid(row=1,column=1,padx=20)
        Label(frame, text="Nombres", font="Arial 9 bold").grid(row=1,column=2,padx=20)
        Label(frame, text="Grado", font="Arial 9 bold").grid(row=1,column=3,padx=20)

        Label(frame, text="Lugar de Nacimiento:", font="Arial 9 bold").grid(row=2,column=0,padx=20,pady=20,sticky="e")
        Label(frame, text=listaMostrar[5],bg="white").grid(row=2,column=1,padx=5,pady=20,sticky="w")
        
        Label(frame, text="Fecha de Nacimiento:", font="Arial 9 bold").grid(row=3,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[6],bg="white").grid(row=3,column=1,padx=5,pady=5, sticky="w")
        Label(frame, text="Escalafón:", font="Arial 9 bold").grid(row=3,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[7], bg="white").grid(row=3,column=3,padx=20,pady=5, sticky="w")
        
        Label(frame, text="C.I:", font="Arial 9 bold").grid(row=4,column=0,padx=20,pady=5,sticky="w")
        Label(frame, text=listaMostrar[0], bg="white").grid(row=4,column=0,padx=0,pady=5)
        Label(frame, text="Serie:", font="Arial 9 bold").grid(row=4,column=1,padx=0,pady=5,sticky="w")
        Label(frame, text=listaMostrar[8], bg="white").grid(row=4,column=1,padx=0,pady=5)
        Label(frame, text="Seccion:", font="Arial 9 bold").grid(row=4,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[9], bg="white").grid(row=4,column=3,padx=20,pady=5, sticky="w")
        Label(frame, text="Otorgado:", font="Arial 9 bold").grid(row=4,column=3,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[10], bg="white").grid(row=4,column=4,padx=0,pady=5, sticky="w")

        Label(frame, text="Estado Civil:", font="Arial 9 bold").grid(row=5,column=0,padx=20,pady=5,sticky="w")
        Label(frame, text=listaMostrar[11], bg="white").grid(row=5,column=0,padx=0,pady=5)        
        Label(frame, text="Profesión:", font="Arial 9 bold").grid(row=5,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[12], bg="white").grid(row=5,column=3,padx=20,pady=5, sticky="w")

        Label(frame, text="Domicilio:", font="Arial 9 bold").grid(row=6,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[13], bg="white").grid(row=6,column=1,padx=0,pady=5, sticky="w")        
        Label(frame, text="Nº:", font="Arial 9 bold").grid(row=6,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[14], bg="white").grid(row=6,column=3,padx=20,pady=5, sticky="w")

        Label(frame, text="Zona:", font="Arial 9 bold").grid(row=7,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[15], bg="white").grid(row=7,column=1,padx=0,pady=5, sticky="w")        
        Label(frame, text="Telefono:", font="Arial 9 bold").grid(row=7,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[16], bg="white").grid(row=7,column=3,padx=20,pady=5, sticky="w")
        Label(frame, text="Celular:", font="Arial 9 bold").grid(row=7,column=3,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[17], bg="white").grid(row=7,column=4,padx=0,pady=5, sticky="w")
        
        Label(frame, text="Libreta de Servicio Militar:", font="Arial 9 bold").grid(row=8,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[18], bg="white").grid(row=8,column=1,padx=0,pady=5, sticky="w")        
        Label(frame, text="Nombre de la Madre:", font="Arial 9 bold").grid(row=9,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[19], bg="white").grid(row=9,column=1,padx=0,pady=5, sticky="w")
        Label(frame, text="Nombre del Padre:", font="Arial 9 bold").grid(row=10,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[20], bg="white").grid(row=10,column=1,padx=0,pady=5, sticky="w")
        Label(frame, text="En caso de emergencia llamar a:", font="Arial 9 bold").grid(row=11,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[21], bg="white").grid(row=11,column=1,padx=0,pady=5, sticky="w")
        
        Label(frame, text="Actual Destino:", font="Arial 9 bold").grid(row=12,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[22], bg="white").grid(row=12,column=1,padx=0,pady=5, sticky="w")
        Label(frame, text="Unidad:", font="Arial 9 bold").grid(row=12,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[23], bg="white").grid(row=12,column=3,padx=0,pady=5, sticky="w")

        Label(frame, text="Fecha de Ingreso a Pol. Bol.:", font="Arial 9 bold").grid(row=13,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[24], bg="white").grid(row=13,column=1,padx=0,pady=5, sticky="w")

        Label(frame, text="Baja:", font="Arial 9 bold").grid(row=14,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[25], bg="white").grid(row=14,column=1,padx=0,pady=5, sticky="w")        
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=14,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[26], bg="white").grid(row=14,column=3,padx=0,pady=5, sticky="w")
        Label(frame, text="Incorporación:", font="Arial 9 bold").grid(row=15,column=0,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[27], bg="white").grid(row=15,column=1,padx=0,pady=5, sticky="w")
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=15,column=2,padx=20,pady=5,sticky="e")
        Label(frame, text=listaMostrar[28], bg="white").grid(row=15,column=3,padx=0,pady=5, sticky="w")

        frame2 = LabelFrame(framePrincipal,text="Otras Referencias", font="Arial 10 bold")
        frame2.grid(row=1,column=0,pady=2)
        frame2.configure(bg="#fff")

        otras_ref = self.otras_referencias(1)[pos]

        Label(frame2, text="Conduce Vehiculo:", font="Arial 9 bold").grid(row=0,column=0,padx=20,pady=5,sticky="e")
        Label(frame2, text=otras_ref[1], bg="white").grid(row=0,column=1,padx=0,pady=5, sticky="w")
        Label(frame2, text="Automóvil:", font="Arial 9 bold").grid(row=1,column=0,padx=20,pady=5,sticky="e")
        Label(frame2, text=otras_ref[2], bg="white").grid(row=1,column=1,padx=0,pady=5, sticky="w")
        Label(frame2, text="Motocicleta:", font="Arial 9 bold").grid(row=2,column=0,padx=20,pady=5,sticky="e")
        Label(frame2, text=otras_ref[3], bg="white").grid(row=2,column=1,padx=0,pady=5, sticky="w")
        Label(frame2, text="Tiene Licencia:", font="Arial 9 bold").grid(row=3,column=0,padx=20,pady=5,sticky="e")
        Label(frame2, text=otras_ref[4], bg="white").grid(row=3,column=1,padx=0,pady=5, sticky="w")
        Label(frame2, text="Otra ocupación u oficio:", font="Arial 9 bold").grid(row=4,column=0,padx=20,pady=5,sticky="e")
        Label(frame2, text=otras_ref[5], bg="white").grid(row=4,column=1,padx=0,pady=5, sticky="w")

        frame3 = LabelFrame(framePrincipal,text="Grupo Familiar", font="Arial 10 bold")
        frame3.grid(row=2,column=0,pady=2)
        frame3.configure(bg="#fff")

        self.tree = ttk.Treeview(frame3,height=10,columns=7)
        self.tree.grid(row=0,column=0,columnspan=5)

        self.tree["columns"] = ("1","2","3","4","5")

        self.tree.column("#0", minwidth=170, width=170,anchor=CENTER)
        self.tree.column("1", minwidth=80, width=80,anchor=CENTER)
        self.tree.column("2", minwidth=90, width=90,anchor=CENTER)
        self.tree.column("3", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("4", minwidth=120, width=120,anchor=CENTER)
        self.tree.column("5", minwidth=120, width=120,anchor=CENTER)

        self.tree.heading("#0",text="Nombres y Apellidos")
        self.tree.heading("1",text="Edad")
        self.tree.heading("2",text="Relación")
        self.tree.heading("3",text="Estado civil")
        self.tree.heading("4",text="Grado Instrucción")
        self.tree.heading("5",text="Ocupación")

        self.cargar_datos_familia(listaMostrar)
        
        ttk.Button(frame3,text="Añadir", command=self.aniadir_Familiares).grid(row=1,column=4,padx=3,pady=3)

        return

    def cargar_datos_familia(self,listaMostrar):
        self.datos_fam = self.otras_referencias(2)
        pos_fam = self.datos_family(self.datos_fam,listaMostrar)
        records = self.tree.get_children()
        for element in records:
            self.tree.delete(element)

        for fila in pos_fam:
            self.tree.insert("","end",text=self.datos_fam[fila][1],
                                values=(self.datos_fam[fila][2],self.datos_fam[fila][3],
                                        self.datos_fam[fila][4],self.datos_fam[fila][5],
                                        self.datos_fam[fila][6]))

    def aniadir_Familiares(self):
        self.aniadirFam = Toplevel(self.newWindow)
        self.aniadirFam.title('Datos Personal')
        self.aniadirFam.geometry("330x250+200+150")

        frame = LabelFrame(self.aniadirFam,text="Datos Personales del Funcionario", font="Arial 10 bold")
        frame.grid(row=0,column=0,padx=2,pady=2)
        frame.configure(bg="#f5f5f5")

        Label(frame, text="Nombres y Apellidos:", font="Arial 9 bold").grid(row=0,column=0,padx=3,pady=5,sticky='e')
        self.nombres = Entry(frame, width=30) 
        self.nombres.focus()
        self.nombres.grid(row=0,column=1,padx=3,pady=5, sticky="w")
        Label(frame, text="Edad:", font="Arial 9 bold").grid(row=1,column=0,padx=3,pady=5, sticky='e')
        self.edad = Entry(frame, width=15) 
        self.edad.grid(row=1,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Relación:", font="Arial 9 bold").grid(row=2,column=0,padx=3,pady=5, sticky='e')
        self.relacion = Entry(frame, width=10) 
        self.relacion.grid(row=2,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Estado Civil:", font="Arial 9 bold").grid(row=3,column=0,padx=3,pady=5, sticky='e')
        self.estadoCivil = Entry(frame, width=10) 
        self.estadoCivil.grid(row=3,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Grado Instrucción:", font="Arial 9 bold").grid(row=4,column=0,padx=3,pady=5, sticky='e')
        self.grado = Entry(frame, width=10) 
        self.grado.grid(row=4,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Ocupación:", font="Arial 9 bold").grid(row=5,column=0,padx=3,pady=5, sticky='e')
        self.ocupacion = Entry(frame, width=15) 
        self.ocupacion.grid(row=5,column=1,padx=3,pady=5,sticky='w')
        ttk.Button(frame,text="Aceptar",command=self.guardar_datos).grid(row=6,column=1,pady=5,padx=3,sticky='w')
        ttk.Button(frame,text="Cancelar",command=self.cancelar).grid(row=6,column=1,pady=5,padx=3,sticky='e')

    def guardar_datos(self):
        self.lista_fam = []
        if (len(self.nombres.get())!=0 and len(self.edad.get())!=0 and len(self.relacion.get())!=0 
            and len(self.estadoCivil.get())!=0 and len(self.grado.get())!=0 and len(self.ocupacion.get())!=0):
            self.lista_fam.append(self.lista_mostrar[0])
            self.lista_fam.append(self.nombres.get())
            self.lista_fam.append(self.edad.get())
            self.lista_fam.append(self.relacion.get())
            self.lista_fam.append(self.estadoCivil.get())
            self.lista_fam.append(self.grado.get())
            self.lista_fam.append(self.ocupacion.get())

            self.lista_fam = tuple(self.lista_fam)

            self.filesheet = "./excel/datos.xlsx"
            # self.filesheet = self.archivo
            wb = load_workbook(self.filesheet)
            nombresHojas = wb.sheetnames
            sheet = wb[nombresHojas[2]]
            self.listaDatos = []
            self.listaDatos.append(self.lista_fam)

            for row in self.listaDatos:
                sheet.append(row)

            wb.save(self.filesheet)
            self.cargar_datos_familia(self.lista_mostrar)
            self.listaDatos.clear()
        else:
            messagebox.showwarning("Advertencia", "Falta llenar datos")
            return
        self.aniadirFam.destroy()

    def cancelar(self):
        self.aniadirFam.destroy()

    def otras_referencias(self,pos):
        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        sheet = wb[nombresHojas[pos]]

        listaDeListas =[]

        for fila in range(2,sheet.max_row+1):
            lista = []
            for columna in range(1,sheet.max_column+1):
                lista.append(sheet.cell(row=fila,column=columna).value)
            listaDeListas.append(lista)
        return listaDeListas

    def datos_family(self, datos_family,listaMostrar):
        cont=-1
        lista_res=[]
        for lista in datos_family:
            cont+=1
            for dato in lista:
                if str(listaMostrar[0]) in str(dato) and len(str(listaMostrar[0])) == len(str(dato)):
                    lista_res.append(cont)
                    break
        return lista_res

class Aniadir:
    def __init__(self,ventana):
        self.newWindow = Toplevel(ventana)
        self.newWindow.title('Datos Personal')
        # self.newWindow.iconbitmap('./img/police2.ico')
        self.newWindow.geometry("1150x510+150+100")

        self.miCanvas = Canvas(self.newWindow,background="white", width=770,height=570)
        framePrincipal = Frame(self.miCanvas)
        miScroll = Scrollbar(self.newWindow, orient="vertical")        

        miScroll.pack(side="right",fill="y")
        miScroll.config(command=self.miCanvas.yview)
        self.miCanvas.configure(yscrollcommand=miScroll.set)
        self.miCanvas.pack(side="left",expand=True,fill="both")
        self.miCanvas.create_window((0,0),window=framePrincipal,anchor="nw")

        self.cargar_ventanas(framePrincipal)
        
        self.newWindow.update()
        self.miCanvas.config(scrollregion=self.miCanvas.bbox("all"))

    def cargar_ventanas(self, framePrincipal):
        frame = LabelFrame(framePrincipal,text="Datos Personales del Funcionario", font="Arial 10 bold")
        frame.grid(row=0,column=0,padx=2,pady=2)
        frame.configure(bg="#f5f5f5")
    
        Label(frame, text="Apellido Paterno:", font="Arial 9 bold").grid(row=0,column=0,padx=3,pady=5,sticky='e')
        self.apellidoPaterno = Entry(frame, width=15) 
        self.apellidoPaterno.focus()
        self.apellidoPaterno.grid(row=0,column=1,padx=3,pady=5, sticky="w")
        Label(frame, text="Apellido Materno:", font="Arial 9 bold").grid(row=0,column=2,padx=3,pady=5, sticky='e')
        self.apellidoMaterno = Entry(frame, width=15) 
        self.apellidoMaterno.grid(row=0,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombres:", font="Arial 9 bold").grid(row=0,column=4,padx=3,pady=5,sticky='e')
        self.nombres = Entry(frame, width=20) 
        self.nombres.grid(row=0,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Grado:", font="Arial 9 bold").grid(row=0,column=6,padx=3,pady=5,sticky='e')
        self.grado = Entry(frame, width=10) 
        self.grado.grid(row=0,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Lugar de Nacimiento:", font="Arial 9 bold").grid(row=1,column=0,padx=3,pady=5)
        self.lugarNac = Entry(frame, width=25) 
        self.lugarNac.grid(row=1,column=1,padx=3,pady=5)
        Label(frame, text="Fecha de Nacimiento:", font="Arial 9 bold").grid(row=1,column=2,padx=3,pady=5)
        self.fechaNac = Entry(frame, width=20) 
        self.fechaNac.grid(row=1,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Escalafón:", font="Arial 9 bold").grid(row=1,column=4,padx=3,pady=5,sticky='e')
        self.escalafon = Entry(frame, width=10) 
        self.escalafon.grid(row=1,column=5,padx=3,pady=5, sticky="w")
        Label(frame, text="C.I.:", font="Arial 9 bold").grid(row=1,column=6,padx=3,pady=5,sticky='e')
        self.ci = Entry(frame, width=10) 
        self.ci.grid(row=1,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Serie:", font="Arial 9 bold").grid(row=2,column=0,padx=3,pady=5,sticky="w")
        self.serie = Entry(frame, width=13) 
        self.serie.grid(row=2,column=0,padx=3,pady=5,sticky="e")
        Label(frame, text="Sección:", font="Arial 9 bold").grid(row=2,column=1,padx=3,pady=5,sticky='w')
        self.seccion = Entry(frame, width=15) 
        self.seccion.grid(row=2,column=1,padx=3,pady=5,sticky='e')
        Label(frame, text="Otorgado:", font="Arial 9 bold").grid(row=2,column=2,padx=3,pady=5,sticky='e')
        self.otorgado = Combobox(frame, state="readonly")
        self.otorgado['values'] = ['La Paz','Oruro','Potosi','Cochabamba','Chuquisaca','Tarija','Pando','Beni','Santa Cruz']
        self.otorgado.grid(row=2,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Estado Civil:", font="Arial 9 bold").grid(row=2,column=4,padx=3,pady=5,sticky='e')
        self.estadoCivil = Entry(frame, width=10) 
        self.estadoCivil.grid(row=2,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Profesión:", font="Arial 9 bold").grid(row=2,column=6,padx=3,pady=5,sticky='e')
        self.profesion = Entry(frame, width=10) 
        self.profesion.grid(row=2,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Domicilio:", font="Arial 9 bold").grid(row=3,column=0,padx=3,pady=5,sticky='e')
        self.domicilio = Entry(frame, width=25) 
        self.domicilio.grid(row=3,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Zona:", font="Arial 9 bold").grid(row=3,column=2,padx=3,pady=5,sticky='e')
        self.zona = Entry(frame, width=25) 
        self.zona.grid(row=3,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nº:", font="Arial 9 bold").grid(row=3,column=4,padx=3,pady=5,sticky='w')
        self.nro = Entry(frame, width=13) 
        self.nro.grid(row=3,column=4,padx=3,pady=5,sticky='e')
        Label(frame, text="Telefono:", font="Arial 9 bold").grid(row=3,column=5,padx=3,pady=5,sticky='w')
        self.telefono = Entry(frame, width=12)
        self.telefono.grid(row=3,column=5,padx=3,pady=5,sticky='e')
        Label(frame, text="Celular:", font="Arial 9 bold").grid(row=3,column=6,padx=3,pady=5,sticky='e')
        self.celular = Entry(frame, width=10)
        self.celular.grid(row=3,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Libreta Serv. Militar:", font="Arial 9 bold").grid(row=4,column=0,padx=3,pady=5,sticky='e')
        self.servMil = Entry(frame, width=25)
        self.servMil.grid(row=4,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombre de la Madre:", font="Arial 9 bold").grid(row=4,column=2,padx=3,pady=5,sticky='e')
        self.nombreMadre = Entry(frame, width=25)
        self.nombreMadre.grid(row=4,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombre del Padre:", font="Arial 9 bold").grid(row=4,column=4,padx=3,pady=5,sticky='e')
        self.nombrePadre = Entry(frame, width=25)
        self.nombrePadre.grid(row=4,column=5,padx=3,pady=5,sticky='w')

        Label(frame, text="Unidad:", font="Arial 9 bold").grid(row=5,column=0,padx=3,pady=5,sticky='e')
        self.unidad = Entry(frame, width=25)
        self.unidad.grid(row=5,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Actual Destino:", font="Arial 9 bold").grid(row=5,column=2,padx=3,pady=5,sticky='e')
        self.destino = Entry(frame, width=25)
        self.destino.grid(row=5,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Emergencia llamar a:", font="Arial 9 bold").grid(row=5,column=4,padx=3,pady=5,sticky='e')
        self.emergencia = Entry(frame, width=25)
        self.emergencia.grid(row=5,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Fecha Ingreso Pol. Bol.:", font="Arial 9 bold").grid(row=5,column=6,padx=3,pady=5,sticky='e')
        self.ingresoPol = Entry(frame, width=15)
        self.ingresoPol.grid(row=5,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Baja:", font="Arial 9 bold").grid(row=6,column=0,padx=3,pady=5,sticky='e')
        self.baja = Entry(frame, width=25)
        self.baja.grid(row=6,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=6,column=2,padx=3,pady=5,sticky='e')
        self.motivo1 = Entry(frame, width=25)
        self.motivo1.grid(row=6,column=3,padx=3,pady=5,sticky='w')
        
        Label(frame, text="Incorporación:", font="Arial 9 bold").grid(row=7,column=0,padx=3,pady=5,sticky='e')
        self.incorporacion = Entry(frame, width=25)
        self.incorporacion.grid(row=7,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=7,column=2,padx=3,pady=5,sticky='e')
        self.motivo2 = Entry(frame, width=25)
        self.motivo2.grid(row=7,column=3,padx=3,pady=5,sticky='w')

        frame2 = LabelFrame(framePrincipal,text="Otras Referencias", font="Arial 10 bold")
        frame2.grid(row=1,column=0,padx=2,pady=2)
        frame2.configure(bg="#f5f5f5")

        self.opcionconduce = StringVar()
        self.vehiculo = StringVar()
        self.moticicleta = StringVar()
    
        Label(frame2, text="Conduce Vehiculo:", font="Arial 9 bold").grid(row=0,column=0,padx=3,pady=5,sticky='e')
        self.conduceSi = Radiobutton(frame2, text="Si", variable=self.opcionconduce,value="Si")
        self.conduceSi.grid(row=0,column=1,padx=3,pady=5,sticky='w')
        self.conduceNo = Radiobutton(frame2, text="No", variable=self.opcionconduce, value="No")
        self.conduceNo.grid(row=0,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Automóvil:", font="Arial 9 bold").grid(row=1,column=0,padx=3,pady=5,sticky='e')
        self.vehiculoSi = Radiobutton(frame2, text="Si", variable=self.vehiculo,value="Si")
        self.vehiculoSi.grid(row=1,column=1,padx=3,pady=5,sticky='w')
        self.vehiculoNo = Radiobutton(frame2, text="No", variable=self.vehiculo, value="No")
        self.vehiculoNo.grid(row=1,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Motocicleta:", font="Arial 9 bold").grid(row=2,column=0,padx=3,pady=5,sticky='e')
        self.vehiculoSi = Radiobutton(frame2, text="Si", variable=self.moticicleta,value="Si")
        self.vehiculoSi.grid(row=2,column=1,padx=3,pady=5,sticky='w')
        self.vehiculoNo = Radiobutton(frame2, text="No", variable=self.moticicleta, value="No    ")
        self.vehiculoNo.grid(row=2,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Tiene Licencia:", font="Arial 9 bold").grid(row=3,column=0,padx=3,pady=5,sticky='e')
        self.licencia = Entry(frame2, width=20)
        self.licencia.grid(row=3,column=1,padx=3,pady=5,sticky='w')
        Label(frame2, text="Otra ocupacion u oficio:", font="Arial 9 bold").grid(row=4,column=0,padx=3,pady=5,sticky='e')
        self.otraOcupacion = Entry(frame2, width=20)
        self.otraOcupacion.grid(row=4,column=1,padx=3,pady=5,sticky='w')

        frame3 = Label(framePrincipal)
        frame3.grid(row=2,column=0,padx=2,pady=2)
        frame3.configure(bg="#f5f5f5")

        ttk.Button(frame3,text="Aceptar", command=self.guardar_datos).grid(row=0,column=10,padx=0,pady=3)
        ttk.Button(frame3,text="Cancelar", command=self.cancelar).grid(row=0,column=11,padx=0,pady=3)

    def guardar_datos(self):
        self.guardar_datos_policia()
        self.guardar_datos_ref()

    def guardar_datos_policia(self):
        self.lista_fam = []
        if (len(self.apellidoPaterno.get())!=0 and len(self.apellidoMaterno.get())!=0 and len(self.nombres.get())!=0 
            and len(self.grado.get())!=0 and len(self.lugarNac.get())!=0 and len(self.fechaNac.get())!=0
            and len(self.escalafon.get())!=0 and len(self.ci.get())!=0 and len(self.serie.get())!=0
            and len(self.seccion.get())!=0 and len(self.estadoCivil.get())!=0 and len(self.otorgado.get())!=0
            and len(self.profesion.get())!=0 and len(self.domicilio.get())!=0 and len(self.zona.get())!=0
            and len(self.nro.get())!=0 and len(self.celular.get())!=0
            and len(self.servMil.get())!=0 and len(self.nombreMadre.get())!=0 and len(self.nombrePadre.get())!=0
            and len(self.unidad.get())!=0 and len(self.destino.get())!=0 and len(self.emergencia.get())!=0
            and len(self.ingresoPol.get())!=0):
            self.lista_fam.append(self.ci.get())
            self.lista_fam.append(self.apellidoPaterno.get())
            self.lista_fam.append(self.apellidoMaterno.get())
            self.lista_fam.append(self.nombres.get())
            self.lista_fam.append(self.grado.get())
            self.lista_fam.append(self.lugarNac.get())
            self.lista_fam.append(self.fechaNac.get())
            self.lista_fam.append(self.escalafon.get())  
            self.lista_fam.append(self.serie.get())  
            self.lista_fam.append(self.seccion.get())  
            self.lista_fam.append(self.otorgado.get())  
            self.lista_fam.append(self.estadoCivil.get())  
            self.lista_fam.append(self.profesion.get())  
            self.lista_fam.append(self.domicilio.get())  
            self.lista_fam.append(self.nro.get())  
            self.lista_fam.append(self.zona.get())  
            self.lista_fam.append(self.telefono.get())  
            self.lista_fam.append(self.celular.get())  
            self.lista_fam.append(self.servMil.get())  
            self.lista_fam.append(self.nombreMadre.get())  
            self.lista_fam.append(self.nombrePadre.get())  
            self.lista_fam.append(self.emergencia.get())  
            self.lista_fam.append(self.destino.get())  
            self.lista_fam.append(self.unidad.get())  
            self.lista_fam.append(self.ingresoPol.get())  
            self.lista_fam.append(self.baja.get())  
            self.lista_fam.append(self.motivo1.get())  
            self.lista_fam.append(self.incorporacion.get())  
            self.lista_fam.append(self.motivo2.get())  

            self.lista_fam = tuple(self.lista_fam)

            self.filesheet = "./excel/datos.xlsx"
            # self.filesheet = self.archivo
            wb = load_workbook(self.filesheet)
            nombresHojas = wb.sheetnames
            sheet = wb[nombresHojas[0]]
            self.listaDatos = []
            self.listaDatos.append(self.lista_fam)

            for row in self.listaDatos:
                sheet.append(row)

            wb.save(self.filesheet)
            self.listaDatos.clear()
            self.newWindow.destroy()
        else:
            messagebox.showwarning("Advertencia", "Falta llenar datos")
            return

    def guardar_datos_ref(self):
        self.lista_ref = []
        self.lista_ref.append(self.ci.get())
        self.lista_ref.append(self.opcionconduce.get())
        self.lista_ref.append(self.vehiculo.get())
        self.lista_ref.append(self.moticicleta.get())
        self.lista_ref.append(self.licencia.get())
        self.lista_ref.append(self.otraOcupacion.get())

        self.lista_ref = tuple(self.lista_ref)

        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        sheet = wb[nombresHojas[1]]
        self.listaDatos = []
        self.listaDatos.append(self.lista_ref)

        for row in self.listaDatos:
            sheet.append(row)

        wb.save(self.filesheet)
        self.listaDatos.clear()

    def cancelar(self):
        self.newWindow.destroy()

class Editar:
    def __init__(self, ventana,listaMostrar, pos):
        self.newWindow = Toplevel(ventana)
        self.newWindow.title('Editar')
        # self.newWindow.iconbitmap('./img/police2.ico')
        self.newWindow.geometry("1150x510+150+100")

        self.miCanvas = Canvas(self.newWindow,background="white", width=770,height=570)
        framePrincipal = Frame(self.miCanvas)
        miScroll = Scrollbar(self.newWindow, orient="vertical")        

        miScroll.pack(side="right",fill="y")
        miScroll.config(command=self.miCanvas.yview)
        self.miCanvas.configure(yscrollcommand=miScroll.set)
        self.miCanvas.pack(side="left",expand=True,fill="both")
        self.miCanvas.create_window((0,0),window=framePrincipal,anchor="nw")

        self.lista = listaMostrar
        self.cargar_ventanas(framePrincipal,listaMostrar, pos)
        
        self.newWindow.update()
        self.miCanvas.config(scrollregion=self.miCanvas.bbox("all"))

    def cargar_ventanas(self, framePrincipal,listaMostrar,pos):
        self.pos = pos
        frame = LabelFrame(framePrincipal,text="Datos Personales del Funcionario", font="Arial 10 bold")
        frame.grid(row=0,column=0,padx=2,pady=2)
        frame.configure(bg="#f5f5f5")
    
        Label(frame, text="Apellido Paterno:", font="Arial 9 bold").grid(row=0,column=0,padx=3,pady=5,sticky='e')
        self.apellidoPaterno = Entry(frame, width=15) 
        self.apellidoPaterno.insert(0,listaMostrar[1])
        self.apellidoPaterno.focus()
        self.apellidoPaterno.grid(row=0,column=1,padx=3,pady=5, sticky="w")
        Label(frame, text="Apellido Materno:", font="Arial 9 bold").grid(row=0,column=2,padx=3,pady=5, sticky='e')
        self.apellidoMaterno = Entry(frame, width=15) 
        self.apellidoMaterno.insert(0,listaMostrar[2])
        self.apellidoMaterno.grid(row=0,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombres:", font="Arial 9 bold").grid(row=0,column=4,padx=3,pady=5,sticky='e')
        self.nombres = Entry(frame, width=20) 
        self.nombres.insert(0,listaMostrar[3])
        self.nombres.grid(row=0,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Grado:", font="Arial 9 bold").grid(row=0,column=6,padx=3,pady=5,sticky='e')
        self.grado = Entry(frame, width=10) 
        self.grado.insert(0,listaMostrar[4])
        self.grado.grid(row=0,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Lugar de Nacimiento:", font="Arial 9 bold").grid(row=1,column=0,padx=3,pady=5)
        self.lugarNac = Entry(frame, width=25) 
        self.lugarNac.insert(0,listaMostrar[5])
        self.lugarNac.grid(row=1,column=1,padx=3,pady=5)
        Label(frame, text="Fecha de Nacimiento:", font="Arial 9 bold").grid(row=1,column=2,padx=3,pady=5)
        self.fechaNac = Entry(frame, width=20) 
        self.fechaNac.insert(0,listaMostrar[6])
        self.fechaNac.grid(row=1,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Escalafón:", font="Arial 9 bold").grid(row=1,column=4,padx=3,pady=5,sticky='e')
        self.escalafon = Entry(frame, width=10) 
        self.escalafon.insert(0,listaMostrar[7])
        self.escalafon.grid(row=1,column=5,padx=3,pady=5, sticky="w")
        Label(frame, text="C.I.:", font="Arial 9 bold").grid(row=1,column=6,padx=3,pady=5,sticky='e')
        self.ci = Entry(frame, width=10) 
        self.ci.insert(0,listaMostrar[0])
        self.ci.grid(row=1,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Serie:", font="Arial 9 bold").grid(row=2,column=0,padx=3,pady=5,sticky="w")
        self.serie = Entry(frame, width=13) 
        self.serie.insert(0,listaMostrar[8])
        self.serie.grid(row=2,column=0,padx=3,pady=5,sticky="e")
        Label(frame, text="Sección:", font="Arial 9 bold").grid(row=2,column=1,padx=3,pady=5,sticky='w')
        self.seccion = Entry(frame, width=15) 
        self.seccion.insert(0,listaMostrar[9])
        self.seccion.grid(row=2,column=1,padx=3,pady=5,sticky='e')
        Label(frame, text="Otorgado:", font="Arial 9 bold").grid(row=2,column=2,padx=3,pady=5,sticky='e')
        self.otorgado = Combobox(frame, state="readonly")
        self.otorgado['values'] = ['La Paz','Oruro','Potosi','Cochabamba','Chuquisaca','Tarija','Pando','Beni','Santa Cruz']
        self.otorgado.grid(row=2,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Estado Civil:", font="Arial 9 bold").grid(row=2,column=4,padx=3,pady=5,sticky='e')
        self.estadoCivil = Entry(frame, width=10) 
        self.estadoCivil.insert(0,listaMostrar[11])
        self.estadoCivil.grid(row=2,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Profesión:", font="Arial 9 bold").grid(row=2,column=6,padx=3,pady=5,sticky='e')
        self.profesion = Entry(frame, width=10) 
        self.profesion.insert(0,listaMostrar[12])
        self.profesion.grid(row=2,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Domicilio:", font="Arial 9 bold").grid(row=3,column=0,padx=3,pady=5,sticky='e')
        self.domicilio = Entry(frame, width=25) 
        self.domicilio.insert(0,listaMostrar[13])
        self.domicilio.grid(row=3,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Zona:", font="Arial 9 bold").grid(row=3,column=2,padx=3,pady=5,sticky='e')
        self.zona = Entry(frame, width=25) 
        self.zona.insert(0,listaMostrar[15])
        self.zona.grid(row=3,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nº:", font="Arial 9 bold").grid(row=3,column=4,padx=3,pady=5,sticky='w')
        self.nro = Entry(frame, width=13) 
        self.nro.insert(0,listaMostrar[14])
        self.nro.grid(row=3,column=4,padx=3,pady=5,sticky='e')
        Label(frame, text="Telefono:", font="Arial 9 bold").grid(row=3,column=5,padx=3,pady=5,sticky='w')
        self.telefono = Entry(frame, width=12)
        self.telefono.insert(0,listaMostrar[16])
        self.telefono.grid(row=3,column=5,padx=3,pady=5,sticky='e')
        Label(frame, text="Celular:", font="Arial 9 bold").grid(row=3,column=6,padx=3,pady=5,sticky='e')
        self.celular = Entry(frame, width=10)
        self.celular.insert(0,listaMostrar[17])
        self.celular.grid(row=3,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Libreta Serv. Militar:", font="Arial 9 bold").grid(row=4,column=0,padx=3,pady=5,sticky='e')
        self.servMil = Entry(frame, width=25)
        self.servMil.insert(0,listaMostrar[18])
        self.servMil.grid(row=4,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombre de la Madre:", font="Arial 9 bold").grid(row=4,column=2,padx=3,pady=5,sticky='e')
        self.nombreMadre = Entry(frame, width=25)
        self.nombreMadre.insert(0,listaMostrar[19])
        self.nombreMadre.grid(row=4,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Nombre del Padre:", font="Arial 9 bold").grid(row=4,column=4,padx=3,pady=5,sticky='e')
        self.nombrePadre = Entry(frame, width=25)
        self.nombrePadre.insert(0,listaMostrar[20])
        self.nombrePadre.grid(row=4,column=5,padx=3,pady=5,sticky='w')

        Label(frame, text="Unidad:", font="Arial 9 bold").grid(row=5,column=0,padx=3,pady=5,sticky='e')
        self.unidad = Entry(frame, width=25)
        self.unidad.insert(0,listaMostrar[23])
        self.unidad.grid(row=5,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Actual Destino:", font="Arial 9 bold").grid(row=5,column=2,padx=3,pady=5,sticky='e')
        self.destino = Entry(frame, width=25)
        self.destino.insert(0,listaMostrar[22])
        self.destino.grid(row=5,column=3,padx=3,pady=5,sticky='w')
        Label(frame, text="Emergencia llamar a:", font="Arial 9 bold").grid(row=5,column=4,padx=3,pady=5,sticky='e')
        self.emergencia = Entry(frame, width=25)
        self.emergencia.insert(0,listaMostrar[21])
        self.emergencia.grid(row=5,column=5,padx=3,pady=5,sticky='w')
        Label(frame, text="Fecha Ingreso Pol. Bol.:", font="Arial 9 bold").grid(row=5,column=6,padx=3,pady=5,sticky='e')
        self.ingresoPol = Entry(frame, width=15)
        self.ingresoPol.insert(0,listaMostrar[24])
        self.ingresoPol.grid(row=5,column=7,padx=3,pady=5,sticky='w')

        Label(frame, text="Baja:", font="Arial 9 bold").grid(row=6,column=0,padx=3,pady=5,sticky='e')
        self.baja = Entry(frame, width=25)
        self.baja.insert(0,listaMostrar[25])
        self.baja.grid(row=6,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=6,column=2,padx=3,pady=5,sticky='e')
        self.motivo1 = Entry(frame, width=25)
        self.motivo1.insert(0,listaMostrar[26])
        self.motivo1.grid(row=6,column=3,padx=3,pady=5,sticky='w')
        
        Label(frame, text="Incorporación:", font="Arial 9 bold").grid(row=7,column=0,padx=3,pady=5,sticky='e')
        self.incorporacion = Entry(frame, width=25)
        self.incorporacion.insert(0,listaMostrar[27])
        self.incorporacion.grid(row=7,column=1,padx=3,pady=5,sticky='w')
        Label(frame, text="Motivo:", font="Arial 9 bold").grid(row=7,column=2,padx=3,pady=5,sticky='e')
        self.motivo2 = Entry(frame, width=25)
        self.motivo2.insert(0,listaMostrar[28])
        self.motivo2.grid(row=7,column=3,padx=3,pady=5,sticky='w')

        frame2 = LabelFrame(framePrincipal,text="Otras Referencias", font="Arial 10 bold")
        frame2.grid(row=1,column=0,padx=2,pady=2)
        frame2.configure(bg="#f5f5f5")

        self.opcionconduce = StringVar()
        self.vehiculo = StringVar()
        self.moticicleta = StringVar()
    
        Label(frame2, text="Conduce Vehiculo:", font="Arial 9 bold").grid(row=0,column=0,padx=3,pady=5,sticky='e')
        self.conduceSi = Radiobutton(frame2, text="Si", variable=self.opcionconduce,value="Si")
        self.conduceSi.grid(row=0,column=1,padx=3,pady=5,sticky='w')
        self.conduceNo = Radiobutton(frame2, text="No", variable=self.opcionconduce, value="No")
        self.conduceNo.grid(row=0,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Automóvil:", font="Arial 9 bold").grid(row=1,column=0,padx=3,pady=5,sticky='e')
        self.vehiculoSi = Radiobutton(frame2, text="Si", variable=self.vehiculo,value="Si")
        self.vehiculoSi.grid(row=1,column=1,padx=3,pady=5,sticky='w')
        self.vehiculoNo = Radiobutton(frame2, text="No", variable=self.vehiculo, value="No")
        self.vehiculoNo.grid(row=1,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Motocicleta:", font="Arial 9 bold").grid(row=2,column=0,padx=3,pady=5,sticky='e')
        self.vehiculoSi = Radiobutton(frame2, text="Si", variable=self.moticicleta,value="Si")
        self.vehiculoSi.grid(row=2,column=1,padx=3,pady=5,sticky='w')
        self.vehiculoNo = Radiobutton(frame2, text="No", variable=self.moticicleta, value="No    ")
        self.vehiculoNo.grid(row=2,column=1,padx=3,pady=5,sticky='e')
        Label(frame2, text="Tiene Licencia:", font="Arial 9 bold").grid(row=3,column=0,padx=3,pady=5,sticky='e')
        self.licencia = Entry(frame2, width=20)
        self.licencia.grid(row=3,column=1,padx=3,pady=5,sticky='w')
        Label(frame2, text="Otra ocupacion u oficio:", font="Arial 9 bold").grid(row=4,column=0,padx=3,pady=5,sticky='e')
        self.otraOcupacion = Entry(frame2, width=20)
        self.otraOcupacion.grid(row=4,column=1,padx=3,pady=5,sticky='w')

        frame3 = Label(framePrincipal)
        frame3.grid(row=2,column=0,padx=2,pady=2)
        frame3.configure(bg="#f5f5f5")

        ttk.Button(frame3,text="Aceptar",command=self.guardar_datos).grid(row=0,column=10,padx=0,pady=3)
        ttk.Button(frame3,text="Cancelar", command=self.cancelar).grid(row=0,column=11,padx=0,pady=3)

    def guardar_datos(self):
        self.guardar_datos_policia()
        self.guardar_datos_ref()
        self.newWindow.destroy()


    def guardar_datos_policia(self):
        self.lista_fam = []
        if (len(self.apellidoPaterno.get())!=0 and len(self.apellidoMaterno.get())!=0 and len(self.nombres.get())!=0 
            and len(self.grado.get())!=0 and len(self.lugarNac.get())!=0 and len(self.fechaNac.get())!=0
            and len(self.escalafon.get())!=0 and len(self.ci.get())!=0 and len(self.serie.get())!=0
            and len(self.seccion.get())!=0 and len(self.estadoCivil.get())!=0 and len(self.otorgado.get())!=0
            and len(self.profesion.get())!=0 and len(self.domicilio.get())!=0 and len(self.zona.get())!=0
            and len(self.nro.get())!=0 and len(self.celular.get())!=0
            and len(self.servMil.get())!=0 and len(self.nombreMadre.get())!=0 and len(self.nombrePadre.get())!=0
            and len(self.unidad.get())!=0 and len(self.destino.get())!=0 and len(self.emergencia.get())!=0
            and len(self.ingresoPol.get())!=0):
            self.lista_fam.append(self.ci.get())
            self.lista_fam.append(self.apellidoPaterno.get())
            self.lista_fam.append(self.apellidoMaterno.get())
            self.lista_fam.append(self.nombres.get())
            self.lista_fam.append(self.grado.get())
            self.lista_fam.append(self.lugarNac.get())
            self.lista_fam.append(self.fechaNac.get())
            self.lista_fam.append(self.escalafon.get())  
            self.lista_fam.append(self.serie.get())  
            self.lista_fam.append(self.seccion.get())  
            self.lista_fam.append(self.otorgado.get())  
            self.lista_fam.append(self.estadoCivil.get())  
            self.lista_fam.append(self.profesion.get())  
            self.lista_fam.append(self.domicilio.get())  
            self.lista_fam.append(self.nro.get())  
            self.lista_fam.append(self.zona.get())  
            self.lista_fam.append(self.telefono.get())  
            self.lista_fam.append(self.celular.get())  
            self.lista_fam.append(self.servMil.get())  
            self.lista_fam.append(self.nombreMadre.get())  
            self.lista_fam.append(self.nombrePadre.get())  
            self.lista_fam.append(self.emergencia.get())  
            self.lista_fam.append(self.destino.get())  
            self.lista_fam.append(self.unidad.get())  
            self.lista_fam.append(self.ingresoPol.get())  
            self.lista_fam.append(self.baja.get())  
            self.lista_fam.append(self.motivo1.get())  
            self.lista_fam.append(self.incorporacion.get())  
            self.lista_fam.append(self.motivo2.get())

            self.lista_fam = tuple(self.lista_fam)

            self.filesheet = "./excel/datos.xlsx"
            # self.filesheet = self.archivo
            wb = load_workbook(self.filesheet)
            nombresHojas = wb.sheetnames
            sheet = wb[nombresHojas[0]]
            self.listaDatos = []
            self.listaDatos.append(self.lista_fam)

            self.res = 0

            for fila in range(2,sheet.max_row+1):
                for columna in range(1,sheet.max_column+1):
                    if str(self.lista[0]) in (str(sheet.cell(row=fila,column=columna).value)):
                        self.res = fila
                        break

            for row in self.listaDatos:
                sheet.append(row)

            sheet.delete_rows(self.res)
            wb.save(self.filesheet)
            self.listaDatos.clear()
        else:
            messagebox.showwarning("Advertencia", "Falta llenar datos")
            return

    def guardar_datos_ref(self):
        self.lista_ref = []
        self.lista_ref.append(self.ci.get())
        self.lista_ref.append(self.opcionconduce.get())
        self.lista_ref.append(self.vehiculo.get())
        self.lista_ref.append(self.moticicleta.get())
        self.lista_ref.append(self.licencia.get())
        self.lista_ref.append(self.otraOcupacion.get())

        miLista = self.otras_referencias(1)[self.pos]

        self.lista_ref = tuple(self.lista_ref)

        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        sheet = wb[nombresHojas[1]]
        self.listaDatos = []
        self.listaDatos.append(self.lista_ref)

        self.res = 0

        for fila in range(2,sheet.max_row+1):
            for columna in range(1,sheet.max_column+1):
                if str(miLista[0]) in (str(sheet.cell(row=fila,column=columna).value)):
                    self.res = fila
                    break

        for row in self.listaDatos:
            sheet.append(row)

        sheet.delete_rows(self.res)
        wb.save(self.filesheet)
        self.listaDatos.clear()

    def otras_referencias(self,pos):
        self.filesheet = "./excel/datos.xlsx"
        # self.filesheet = self.archivo
        wb = load_workbook(self.filesheet)
        nombresHojas = wb.sheetnames
        sheet = wb[nombresHojas[pos]]

        listaDeListas =[]

        for fila in range(2,sheet.max_row+1):
            lista = []
            for columna in range(1,sheet.max_column+1):
                lista.append(sheet.cell(row=fila,column=columna).value)
            listaDeListas.append(lista)
        return listaDeListas

    def cancelar(self):
        self.newWindow.destroy()


if __name__ == '__main__':
    root = Tk()
    app = Interfaz(root)
    root.mainloop()